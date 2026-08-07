#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Analyze the SKU management Excel file with formulas."""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl

EXCEL_PATH = r'C:\Users\Administrator\Desktop\sku管理带公式版本.xlsx'

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)

for sname in wb.sheetnames:
    ws = wb[sname]
    print(f'=== Sheet: {sname} ===')
    print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')

    # Print all headers with column letter
    print('\n--- ALL HEADERS (col_letter | header_name) ---')
    for cell in ws[1]:
        if cell.value is not None:
            print(f'  {cell.coordinate} | {cell.value}')

    # Print first 2 data rows with ALL values (both formulas and static)
    for row_idx in range(2, min(5, ws.max_row+1)):
        print(f'\n--- Row {row_idx} (DATA) ---')
        for cell in ws[row_idx]:
            val = cell.value
            if val is not None:
                col_letter = cell.coordinate.replace(str(row_idx), '')
                header = ws[f'{col_letter}1'].value or '?'
                is_formula = str(val).startswith('=')
                print(f'  [{cell.coordinate}] {header}: {val}{" [FORMULA]" if is_formula else ""}')

    # Count how many formulas per column (check first 10 rows)
    print('\n--- FORMULA COUNT by column (first 10 data rows) ---')
    for col in range(1, ws.max_column+1):
        col_letter = openpyxl.utils.get_column_letter(col)
        header = ws[f'{col_letter}1'].value or '?'
        formula_count = 0
        formula_samples = []
        for row in range(2, min(12, ws.max_row+1)):
            cell = ws[f'{col_letter}{row}']
            if cell.value and str(cell.value).startswith('='):
                formula_count += 1
                if len(formula_samples) < 2:
                    formula_samples.append(cell.value)
        if formula_count > 0:
            print(f'  [{col_letter}] {header}: {formula_count}/10 formulas. Samples: {formula_samples}')
