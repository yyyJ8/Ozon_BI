"""
OZON直发信息 — 数据导入脚本
从桌面 Excel 导入数据到数据库

用法:
    1. 先激活虚拟环境
    2. cd d:/OzonSku
    3. python scripts/import_ozon_direct.py

功能:
    - 读取 OZON直发信息.xlsx 的 2 张 Sheet
    - 插入 ozon_direct_sku 和 ozon_direct_shipment 表
    - 扫描标签文件和入库清单目录，注册到 ozon_direct_files 表
"""
import os
import shutil
import sys
from datetime import date, datetime

# 确保能找到 app 包
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from app.database import SessionLocal, engine
from app.models import Base, OzonDirectSku, OzonDirectShipment, OzonDirectFile

# 路径配置
DESKTOP = r"C:\Users\Administrator\Desktop\OZON直发信息"
EXCEL_PATH = os.path.join(DESKTOP, "OZON直发信息.xlsx")
LABEL_DIR = os.path.join(DESKTOP, "OZON直发信息-FILE", "SKU基础数据", "标签文件")
RECEIPT_DIR = os.path.join(DESKTOP, "OZON直发信息-FILE", "直发跟进表-N", "入库清单")

# 上传目录
UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "..", "app", "static", "uploads", "ozon_direct")


def ensure_tables():
    """确保新表已创建"""
    Base.metadata.create_all(bind=engine, tables=[
        OzonDirectSku.__table__,
        OzonDirectShipment.__table__,
        OzonDirectFile.__table__,
    ])
    print("OK: tables ready")


def import_sku_data(db):
    """导入 SKU基础数据"""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    if "SKU基础数据" not in wb.sheetnames:
        print("ERR: 未找到 Sheet 'SKU基础数据'")
        return 0
    ws = wb["SKU基础数据"]

    # 清空旧数据
    db.query(OzonDirectFile).filter_by(source_table="sku").delete()
    db.query(OzonDirectSku).delete()
    db.flush()

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        sku_val = str(row[0]).strip() if row[0] else None
        if not sku_val:
            continue
        item = OzonDirectSku(
            sku=sku_val,
            product_name=str(row[1]).strip() if row[1] else None,
            supplier=str(row[2]).strip() if row[2] else None,
            store_name=str(row[3]).strip() if row[3] else None,
            label_file=str(row[4]).strip() if row[4] else None,
        )
        db.add(item)
        count += 1

    db.commit()
    print(f"OK: SKU基础数据: 导入 {count} 条")
    wb.close()
    return count


def import_shipment_data(db):
    """导入 直发跟进表-N"""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    if "直发跟进表-N" not in wb.sheetnames:
        print("ERR: 未找到 Sheet '直发跟进表-N'")
        return 0
    ws = wb["直发跟进表-N"]

    # 清空旧数据
    db.query(OzonDirectFile).filter_by(source_table="shipment").delete()
    db.query(OzonDirectShipment).delete()
    db.flush()

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] and not row[1]:
            continue

        def _str(v):
            return str(v).strip() if v is not None else None

        def _int(v):
            try:
                return int(v) if v is not None else None
            except (ValueError, TypeError):
                return None

        def _date(v):
            if v is None:
                return None
            if isinstance(v, datetime):
                return v.date()
            if isinstance(v, date):
                return v
            if isinstance(v, str):
                v = v.strip()
                for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"]:
                    try:
                        return datetime.strptime(v, fmt).date()
                    except ValueError:
                        pass
            return None

        def _dec(v):
            try:
                return float(v) if v is not None else None
            except (ValueError, TypeError):
                return None

        item = OzonDirectShipment(
            pr_no=_str(row[0]),
            sku=_str(row[1]),
            product_cn_name=_str(row[2]),
            pr_date=_date(row[3]),
            pr_person=_str(row[4]),
            supplier=_str(row[5]),
            po_no=_str(row[6]),
            online_po_no=_str(row[7]),
            is_received=_str(row[8]),
            total_qty=_int(row[9]),
            total_boxes=_int(row[10]),
            product_label=_str(row[11]),
            carton_mark=_str(row[12]),
            warehouse_receipt=_str(row[14]),
            receiving_address=_str(row[16]),
            labeling_notes=_str(row[17]),
            logistics_provider=_str(row[18]),
            first_leg_tracking=_str(row[19]),
            total_boxes_2=_int(row[20]),
            length_cm=_dec(row[21]),
            width_cm=_dec(row[22]),
            height_cm=_dec(row[23]),
            gross_weight=_dec(row[24]),
            total_cbm=_dec(row[25]),
            density=_dec(row[26]),
            plan_no=_str(row[27]),
            ship_date=_date(row[28]),
            tracking_no=_str(row[29]),
            logistics_company=_str(row[30]),
            special_notes=_str(row[31]),
            previous_aftersales=_str(row[32]),
            qty_total_2=_int(row[33]),
            receiving_status=_str(row[34]),
            shipment_no=_str(row[35]),
        )
        db.add(item)
        count += 1

    db.commit()
    print(f"OK: 直发跟进表: 导入 {count} 条")
    wb.close()
    return count


def import_files(db):
    """复制文件到 uploads 目录，注册到数据库"""
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    db.query(OzonDirectFile).delete()
    db.commit()

    total = 0

    # 1. 标签文件
    if os.path.isdir(LABEL_DIR):
        # 获取 SKU 记录，建立文件名到 id 的映射
        skus = db.query(OzonDirectSku).all()
        label_to_sku = {}
        for s in skus:
            if s.label_file:
                label_to_sku[s.label_file.strip()] = s.id

        for fname in os.listdir(LABEL_DIR):
            src = os.path.join(LABEL_DIR, fname)
            if not os.path.isfile(src):
                continue
            ext = os.path.splitext(fname)[1].lower()
            saved_name = f"label_{fname}"
            dst = os.path.join(UPLOAD_DIR, saved_name)
            shutil.copy2(src, dst)

            source_id = label_to_sku.get(fname, 0)
            record = OzonDirectFile(
                source_table="sku",
                source_id=source_id,
                file_name=fname,
                file_path=saved_name,
                file_size=os.path.getsize(dst),
                file_type=ext.lstrip(".") if ext else None,
            )
            db.add(record)
            total += 1

    print(f"  标签文件: {total} 个")

    # 2. 入库清单
    receipt_count = 0
    if os.path.isdir(RECEIPT_DIR):
        shipments = db.query(OzonDirectShipment).all()
        receipt_to_ship = {}
        for s in shipments:
            if s.warehouse_receipt:
                receipt_to_ship[s.warehouse_receipt.strip()] = s.id

        for fname in os.listdir(RECEIPT_DIR):
            src = os.path.join(RECEIPT_DIR, fname)
            if not os.path.isfile(src):
                continue
            ext = os.path.splitext(fname)[1].lower()
            saved_name = f"receipt_{fname}"
            dst = os.path.join(UPLOAD_DIR, saved_name)
            shutil.copy2(src, dst)

            source_id = receipt_to_ship.get(fname, 0)
            record = OzonDirectFile(
                source_table="shipment",
                source_id=source_id,
                file_name=fname,
                file_path=saved_name,
                file_size=os.path.getsize(dst),
                file_type=ext.lstrip(".") if ext else None,
            )
            db.add(record)
            receipt_count += 1

    print(f"  入库清单: {receipt_count} 个")
    total += receipt_count

    db.commit()
    print(f"OK: 文件导入: 共 {total} 个")

    return total


def main():
    print("=" * 50)
    print("OZON直发信息 — 数据导入")
    print("=" * 50)

    if not os.path.exists(EXCEL_PATH):
        print(f"ERR: Excel 文件不存在: {EXCEL_PATH}")
        sys.exit(1)

    ensure_tables()

    db = SessionLocal()
    try:
        n1 = import_sku_data(db)
        n2 = import_shipment_data(db)
        n3 = import_files(db)
        print("=" * 50)
        print(f"完成！SKU: {n1}, 发货: {n2}, 文件: {n3}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
