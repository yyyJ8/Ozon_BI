"""导入 货件汇总.xlsx 到 cargo_shipments 表"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from datetime import datetime
from decimal import Decimal

import openpyxl
from sqlalchemy import text

from app.database import SessionLocal, engine
from app.models import Base, CargoShipment

EXCEL_PATH = r"C:\Users\Administrator\Desktop\货件汇总.xlsx"


def parse_num(v) -> Decimal | None:
    """安全解析数值"""
    if v is None:
        return None
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    s = str(v).strip()
    if not s:
        return None
    try:
        return Decimal(s)
    except Exception:
        return None


def parse_datetime(v) -> datetime | None:
    """安全解析日期时间"""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    if not s:
        return None
    for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"]:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def main():
    # ── 1. 建表 ──
    print("[CREATE] Creating table cargo_shipments ...")
    Base.metadata.create_all(engine)

    # ── 2. 读取 Excel ──
    print(f"[READ] Reading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[wb.sheetnames[0]]

    headers = [str(ws.cell(1, c).value) for c in range(1, 27)]
    print(f"   表头: {headers}")

    # ── 3. 解析行 ──
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        vals = {}
        for col_idx in range(1, 27):
            vals[col_idx] = ws.cell(row_idx, col_idx).value
        rows.append(vals)

    print(f"   共 {len(rows)} 行数据")

    # ── 4. 写入数据库 ──
    db = SessionLocal()
    try:
        # 先清空已有数据
        deleted = db.execute(text("DELETE FROM public.cargo_shipments")).rowcount
        if deleted:
            print(f"[CLEAR] Deleted {deleted} old records")

        count = 0
        for vals in rows:
            shipment = CargoShipment(
                sku=str(vals[1] or "").strip(),
                product_name=str(vals[3] or "").strip() if vals[3] else None,
                store=str(vals[4] or "").strip() if vals[4] else None,
                requisitioner=str(vals[5] or "").strip() if vals[5] else None,
                replenishment_qty=parse_num(vals[6]),
                carton_qty=parse_num(vals[7]),
                carton_volume=parse_num(vals[8]),
                carton_gross_weight=parse_num(vals[9]),
                weight=parse_num(vals[10]),
                cbm=parse_num(vals[11]),
                density=parse_num(vals[12]),
                box_count=parse_num(vals[13]),
                transit_warehouse=str(vals[14] or "").strip() if vals[14] else None,
                logistics_inbound_no=str(vals[15] or "").strip() if vals[15] else None,
                cargo_status=str(vals[16] or "").strip() if vals[16] else None,
                fbo_warehouse_name=str(vals[17] or "").strip() if vals[17] else None,
                booking_code=str(vals[18] or "").strip() if vals[18] else None,
                fbo_listing_time=parse_datetime(vals[19]),
                warehouse_rent_start=parse_datetime(vals[20]),
                actual_listing_qty=parse_num(vals[21]),
                info_remarks=str(vals[22] or "").strip() if vals[22] else None,
                batch_quotation=str(vals[23] or "").strip() if vals[23] else None,
                product_status=str(vals[24] or "").strip() if vals[24] else None,
                stocking_opinion=str(vals[25] or "").strip() if vals[25] else None,
                parent_record=str(vals[26] or "").strip() if vals[26] else None,
            )
            db.add(shipment)
            count += 1

        db.commit()
        print(f"[OK] Import success: {count} records")

    except Exception as e:
        db.rollback()
        print(f"[ERROR] Import failed: {e}")
        raise
    finally:
        db.close()


if __name__ == "__main__":
    main()
