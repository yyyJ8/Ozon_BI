#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Test the SKU formula engine against Excel cached formula values."""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl

from app.services.sku_formulas import compute_formulas, INPUT_FIELDS, COMPUTED_FIELDS

EXCEL_PATH = r'C:\Users\Administrator\Desktop\sku管理带公式版本.xlsx'

# Column mapping: Excel column letter -> field name in DB
COL_MAP = {
    'I': 'length_cm',         'J': 'width_cm',
    'K': 'height_cm',         'L': 'actual_weight_kg',
    'M': 'volume_cbm',        'N': 'density',
    'O': 'first_leg_unit_price', 'P': 'units_per_carton',
    'Q': 'carton_length_cm',  'R': 'carton_width_cm',
    'S': 'carton_height_cm',  'T': 'gross_weight_kg',
    'U': 'volume_liters',     'V': 'purchase_cost_rmb',
    'W': 'purchase_cost_pct', 'X': 'warehousing_fee_rmb',
    'Y': 'fbo_delivery_fee_rmb', 'Z': 'first_leg_cost_rmb',
    'AA': 'first_leg_pct',    'AB': 'acquiring_fee_pct',
    'AC': 'fbo_commission_pct', 'AD': 'logistics_rub',
    'AE': 'delivery_pickup_rub', 'AF': 'last_mile_pct',
    'AG': 'advertising_rate_pct', 'AH': 'return_rate_pct',
    'AI': 'product_cost_rmb', 'AJ': 'exchange_rate',
    'AK': 'price',            'AL': 'green_price_rub',
    'AM': 'discount_pct',     'AN': 'platform_payout_rub',
    'AO': 'actual_payout_rub','AP': 'tax_and_fee_pct',
    'AQ': 'risk_reserve_rub', 'AR': 'profit_rmb',
    'AS': 'profit_rub',       'AT': 'profit_margin_pct',
    'AU': 'competitor_1',     'AV': 'competitor_2',
    'AW': 'competitor_sales',
}


def parse_pct(val):
    """Parse percentage string to decimal, e.g. '2%' -> 0.02, '43%' -> 0.43"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace('%', '').strip()
    try:
        return float(s) / 100
    except ValueError:
        return None


def extract_inputs(ws, row_idx):
    """Extract input fields from a row."""
    inputs = {}
    price = None

    for col_letter, field_name in COL_MAP.items():
        cell = ws[f'{col_letter}{row_idx}']
        val = cell.value

        if val is None:
            continue

        # Skip formula cells for input extraction
        if isinstance(val, str) and val.startswith('='):
            continue

        if field_name == 'price':
            price = float(val)
            continue

        if field_name in INPUT_FIELDS:
            if field_name in ('acquiring_fee_pct', 'fbo_commission_pct',
                              'advertising_rate_pct', 'return_rate_pct', 'purchase_cost_pct'):
                val = parse_pct(val)
            elif isinstance(val, str):
                try:
                    val = float(val)
                except ValueError:
                    pass
            elif isinstance(val, (int, float)):
                val = float(val)
            inputs[field_name] = val

    return inputs, price


def extract_expected(ws_cached, row_idx):
    """Extract expected computed values from cached formulas (data_only=True)."""
    expected = {}
    for col_letter, field_name in COL_MAP.items():
        if field_name not in COMPUTED_FIELDS:
            continue
        cell = ws_cached[f'{col_letter}{row_idx}']
        val = cell.value
        if val is not None:
            if isinstance(val, str):
                val = parse_pct(val)
                if val is None:
                    continue
            expected[field_name] = float(val)
    return expected


def main():
    # Read with formulas (to detect which cells are formulas vs static)
    wb_formula = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
    # Read with cached values (to get formula results)
    wb_cached = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    ws = wb_formula['FBO-SKU信息']
    ws_cached = wb_cached['FBO-SKU信息']

    total = 0
    passed = 0
    failed = 0

    for row_idx in range(2, ws.max_row + 1):
        inputs, price = extract_inputs(ws, row_idx)
        if price is None:
            continue

        expected = extract_expected(ws_cached, row_idx)
        if not expected:
            # No expected values to compare (all formulas and no cached values)
            # Skip this row for validation
            print(f'⏭️  Row {row_idx} [{ws[f"A{row_idx}"].value}] — no cached values to compare')
            continue

        total += 1

        computed = compute_formulas(inputs, price)

        row_errors = []
        for field_name, excel_val in expected.items():
            computed_val = computed.get(field_name)
            if computed_val is None:
                row_errors.append(f'  {field_name}: computed=None, excel={excel_val:.6f}')
                continue

            # Use relative tolerance: 1% for financial calcs, 0.1% for others
            tolerance = max(abs(excel_val) * 0.01, 0.01)
            if abs(computed_val - excel_val) > tolerance:
                row_errors.append(
                    f'  {field_name}: computed={computed_val:.6f}, excel={excel_val:.6f}, '
                    f'diff={computed_val - excel_val:.6f}'
                )

        sku = ws[f'A{row_idx}'].value
        if row_errors:
            failed += 1
            print(f'\n❌ Row {row_idx} [{sku}] ({len(row_errors)}/{len(expected)} mismatch):')
            for err in row_errors:
                print(err)
        else:
            passed += 1
            print(f'✅ Row {row_idx} [{sku}] — {len(expected)} computed fields match')

    print(f'\n{"="*60}')
    print(f'Results: {passed} passed, {failed} failed, {total} total (1% tolerance)')

    # Also print a detailed trace for Row 2 to manually verify
    print(f'\n{"="*60}')
    print('Detailed trace for Row 2 (first data row):')
    inputs, price = extract_inputs(ws, 2)
    computed = compute_formulas(inputs, price)
    print(f'  Inputs: length={inputs.get("length_cm")}, width={inputs.get("width_cm")}, '
          f'height={inputs.get("height_cm")}, weight={inputs.get("actual_weight_kg")}')
    print(f'  units_per_carton={inputs.get("units_per_carton")}, '
          f'first_leg_unit_price={inputs.get("first_leg_unit_price")}')
    print(f'  carton={inputs.get("carton_length_cm")}x{inputs.get("carton_width_cm")}x{inputs.get("carton_height_cm")}')
    print(f'  purchase_cost={inputs.get("purchase_cost_rmb")}, '
          f'product_cost={inputs.get("product_cost_rmb")}, exchange={inputs.get("exchange_rate")}')
    print(f'  green_price={inputs.get("green_price_rub")}, price={price}')
    print(f'  acquiring={inputs.get("acquiring_fee_pct")}, commission={inputs.get("fbo_commission_pct")}')
    print(f'  delivery_pickup={inputs.get("delivery_pickup_rub")}, '
          f'ad={inputs.get("advertising_rate_pct")}, return={inputs.get("return_rate_pct")}')
    print(f'  Computed:')
    for k, v in computed.items():
        print(f'    {k}: {v}')

    return failed == 0


if __name__ == '__main__':
    success = main()
    sys.exit(0 if success else 1)
