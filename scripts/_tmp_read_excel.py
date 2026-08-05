import openpyxl
PATH = r'C:\Users\Administrator\Desktop\新的需求.xlsx'
wb = openpyxl.load_workbook(PATH)
print(f'File: {PATH}')
print(f'Sheets: {wb.sheetnames}')
for sname in wb.sheetnames:
    ws = wb[sname]
    print(f'\n=== {sname}: {ws.max_row} rows x {ws.max_column} cols ===')
    # Print header row (row 1) - column positions and values
    print('Row 1 (headers):')
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is not None:
            print(f'  Col {c}: {v}')
    # Print first data row (row 2)
    print('Row 2 (first data):')
    for c in range(1, ws.max_column + 1):
        v = ws.cell(2, c).value
        if v is not None:
            print(f'  Col {c}: {v}')
    # Print last row
    print(f'Row {ws.max_row} (last):')
    for c in range(1, ws.max_column + 1):
        v = ws.cell(ws.max_row, c).value
        if v is not None:
            print(f'  Col {c}: {v}')
wb.close()
