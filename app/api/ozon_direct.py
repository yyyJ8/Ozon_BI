"""OZON 直发信息 API — SKU 基础数据 / 直发跟进表 / 文件 / 导入导出"""
import os
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import OzonDirectSku, OzonDirectShipment, OzonDirectFile
from app.ozon_direct_logger import log_operation
from app.schemas.ozon_direct import (
    DirectSkuItem, DirectSkuCreate, DirectSkuUpdate,
    DirectShipmentItem, DirectShipmentCreate, DirectShipmentUpdate,
    DirectFileItem,
    PaginatedResponse,
)

router = APIRouter(prefix="/ozon-direct", tags=["ozon-direct"])


# ============================================================
# SKU 基础数据 CRUD
# ============================================================

@router.get("/sku")
def list_sku(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=0, ge=0, le=5000, description="0=不分页返回全部"),
    search: Optional[str] = Query(default=None, description="搜索 SKU/产品名/供应商"),
    db: Session = Depends(get_db),
):
    """SKU 列表（分页+搜索，page_size=0 返回全部）"""
    q = db.query(OzonDirectSku).filter_by(is_deleted=False)
    if search:
        like = f"%{search}%"
        q = q.filter(
            (OzonDirectSku.sku.ilike(like))
            | (OzonDirectSku.product_name.ilike(like))
            | (OzonDirectSku.supplier.ilike(like))
        )
    q = q.order_by(OzonDirectSku.id)
    total = q.count()
    if page_size > 0:
        items = q.offset((page - 1) * page_size).limit(page_size).all()
    else:
        items = q.all()
    return PaginatedResponse(
        items=[DirectSkuItem.model_validate(it) for it in items],
        total=total, page=page, page_size=page_size if page_size > 0 else total,
    )


@router.get("/sku/{sku_id}", response_model=DirectSkuItem)
def get_sku(sku_id: int, db: Session = Depends(get_db)):
    """SKU 详情"""
    item = db.query(OzonDirectSku).filter_by(id=sku_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="SKU 不存在")
    return item


@router.post("/sku", response_model=DirectSkuItem)
def create_sku(body: DirectSkuCreate, db: Session = Depends(get_db)):
    """新增 SKU"""
    item = OzonDirectSku(**body.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    log_operation("CREATE SKU", f"id={item.id} sku={item.sku}")
    return item


@router.put("/sku/{sku_id}", response_model=DirectSkuItem)
def update_sku(sku_id: int, body: DirectSkuUpdate, db: Session = Depends(get_db)):
    """更新 SKU"""
    item = db.query(OzonDirectSku).filter_by(id=sku_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="SKU 不存在")
    for key, val in body.model_dump(exclude_unset=True).items():
        setattr(item, key, val)
    item.updated_at = datetime.now()
    db.commit()
    db.refresh(item)
    log_operation("UPDATE SKU", f"id={sku_id} fields={list(body.model_dump(exclude_unset=True).keys())}")
    return item


@router.delete("/sku/{sku_id}")
def delete_sku(sku_id: int, db: Session = Depends(get_db)):
    """软删除 SKU（前端不展示，数据库保留）"""
    item = db.query(OzonDirectSku).filter_by(id=sku_id, is_deleted=False).first()
    if not item:
        raise HTTPException(status_code=404, detail="SKU 不存在")
    item.is_deleted = True
    item.updated_at = datetime.now()
    db.commit()
    log_operation("DELETE SKU", f"id={sku_id} sku={item.sku}")
    return {"ok": True, "detail": "已删除"}


# ============================================================
# 直发跟进表 CRUD
# ============================================================

@router.get("/shipment")
def list_shipment(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=0, ge=0, le=5000, description="0=不分页返回全部"),
    search: Optional[str] = Query(default=None, description="搜索申购单号/SKU/产品名/供应商"),
    date_from: Optional[date] = Query(default=None, description="申购时间起始"),
    date_to: Optional[date] = Query(default=None, description="申购时间截止"),
    receiving_status: Optional[str] = Query(default=None, description="货物收货情况: 已收到/异常/已取消"),
    db: Session = Depends(get_db),
):
    """发货列表（分页+搜索+日期筛选+收货状态，page_size=0 返回全部）"""
    q = db.query(OzonDirectShipment).filter_by(is_deleted=False)
    if search:
        like = f"%{search}%"
        q = q.filter(
            (OzonDirectShipment.pr_no.ilike(like))
            | (OzonDirectShipment.sku.ilike(like))
            | (OzonDirectShipment.product_cn_name.ilike(like))
            | (OzonDirectShipment.supplier.ilike(like))
        )
    if date_from:
        q = q.filter(OzonDirectShipment.pr_date >= date_from)
    if date_to:
        q = q.filter(OzonDirectShipment.pr_date <= date_to)
    if receiving_status:
        q = q.filter(OzonDirectShipment.receiving_status == receiving_status)
    q = q.order_by(OzonDirectShipment.id)
    total = q.count()
    if page_size > 0:
        items = q.offset((page - 1) * page_size).limit(page_size).all()
    else:
        items = q.all()
    return PaginatedResponse(
        items=[DirectShipmentItem.model_validate(it) for it in items],
        total=total, page=page, page_size=page_size if page_size > 0 else total,
    )


@router.get("/shipment/{ship_id}", response_model=DirectShipmentItem)
def get_shipment(ship_id: int, db: Session = Depends(get_db)):
    """发货详情"""
    item = db.query(OzonDirectShipment).filter_by(id=ship_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="记录不存在")
    return item


@router.post("/shipment", response_model=DirectShipmentItem)
def create_shipment(body: DirectShipmentCreate, db: Session = Depends(get_db)):
    """新增发货记录"""
    item = OzonDirectShipment(**body.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    log_operation("CREATE SHIPMENT", f"id={item.id} pr_no={item.pr_no}")
    return item


@router.put("/shipment/{ship_id}", response_model=DirectShipmentItem)
def update_shipment(ship_id: int, body: DirectShipmentUpdate, db: Session = Depends(get_db)):
    """更新发货记录"""
    item = db.query(OzonDirectShipment).filter_by(id=ship_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="记录不存在")
    for key, val in body.model_dump(exclude_unset=True).items():
        setattr(item, key, val)
    item.updated_at = datetime.now()
    db.commit()
    db.refresh(item)
    log_operation("UPDATE SHIPMENT", f"id={ship_id} fields={list(body.model_dump(exclude_unset=True).keys())}")
    return item


@router.delete("/shipment/{ship_id}")
def delete_shipment(ship_id: int, db: Session = Depends(get_db)):
    """软删除发货记录（前端不展示，数据库保留）"""
    item = db.query(OzonDirectShipment).filter_by(id=ship_id, is_deleted=False).first()
    if not item:
        raise HTTPException(status_code=404, detail="记录不存在")
    item.is_deleted = True
    item.updated_at = datetime.now()
    db.commit()
    log_operation("DELETE SHIPMENT", f"id={ship_id} pr_no={item.pr_no}")
    return {"ok": True, "detail": "已删除"}


# ============================================================
# 文件上传 / 下载 / 删除
# ============================================================

@router.post("/files/upload", response_model=DirectFileItem)
def upload_file(
    file: UploadFile = File(...),
    source_table: str = Query(default="shipment", description="来源表: sku / shipment"),
    source_id: int = Query(default=0, description="来源记录 ID，0 表示暂不关联"),
    db: Session = Depends(get_db),
):
    """上传文件（存入数据库）"""
    content = file.file.read()
    ext = os.path.splitext(file.filename or "file")[1].lower()

    record = OzonDirectFile(
        source_table=source_table,
        source_id=source_id,
        file_name=file.filename or "unknown",
        file_data=content,
        file_size=len(content),
        file_type=ext.lstrip(".") if ext else None,
    )
    db.add(record)
    db.commit()
    db.refresh(record)
    log_operation("UPLOAD FILE", f"id={record.id} name={record.file_name} source={source_table}:{source_id}")
    return record


@router.get("/files/by-source")
def list_files(
    source_table: str = Query(...),
    source_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """查询某条记录的关联文件列表"""
    records = (
        db.query(OzonDirectFile)
        .filter_by(source_table=source_table, source_id=source_id)
        .order_by(OzonDirectFile.uploaded_at.desc())
        .all()
    )
    return [DirectFileItem.model_validate(r) for r in records]


@router.get("/files/{file_id}")
def download_file(file_id: int, db: Session = Depends(get_db)):
    """下载/预览文件"""
    record = db.query(OzonDirectFile).filter_by(id=file_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="文件不存在")
    if not record.file_data:
        raise HTTPException(status_code=404, detail="文件内容为空")
    import io
    return StreamingResponse(
        io.BytesIO(record.file_data),
        media_type="application/octet-stream",
        headers={"Content-Disposition": f"inline; filename={record.file_name}"},
    )


@router.delete("/files/{file_id}")
def delete_file(file_id: int, db: Session = Depends(get_db)):
    """删除文件"""
    record = db.query(OzonDirectFile).filter_by(id=file_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="文件不存在")
    fid = record.id
    fname = record.file_name
    db.delete(record)
    db.commit()
    log_operation("DELETE FILE", f"id={fid} name={fname}")
    return {"ok": True, "detail": "已删除"}


# ============================================================
# Excel 导入
# ============================================================

@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """从 Excel 导入数据（覆盖现有数据）"""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="服务器未安装 openpyxl")

    content = file.file.read()
    import io
    wb = openpyxl.load_workbook(io.BytesIO(content))

    result = {"sku_count": 0, "shipment_count": 0, "errors": []}

    # Sheet 1: SKU基础数据
    if "SKU基础数据" in wb.sheetnames:
        ws = wb["SKU基础数据"]
        # 清空旧数据
        db.query(OzonDirectFile).filter_by(source_table="sku").delete()
        db.query(OzonDirectSku).delete()
        db.flush()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            sku_val = str(row[0]).strip() if row[0] else None
            if not sku_val:
                continue
            item = OzonDirectSku(
                sku=sku_val,
                product_name=str(row[1]).strip() if row[1] else None,
                supplier=str(row[2]).strip() if row[2] else None,
                store_name=str(row[3]).strip() if row[3] else None,
                label_file=str(row[4]).strip() if row[4] else None,
            )
            db.add(item)
        result["sku_count"] = ws.max_row - 1

    # Sheet 2: 直发跟进表-N
    if "直发跟进表-N" in wb.sheetnames:
        ws = wb["直发跟进表-N"]
        db.query(OzonDirectFile).filter_by(source_table="shipment").delete()
        db.query(OzonDirectShipment).delete()
        db.flush()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0] and not row[1]:
                continue

            def _str(v):
                return str(v).strip() if v is not None else None

            def _int(v):
                try:
                    return int(v) if v is not None else None
                except (ValueError, TypeError):
                    return None

            def _date(v):
                if v is None:
                    return None
                if isinstance(v, (date, datetime)):
                    return v.date() if isinstance(v, datetime) else v
                return None

            def _dec(v):
                try:
                    return float(v) if v is not None else None
                except (ValueError, TypeError):
                    return None

            item = OzonDirectShipment(
                pr_no=_str(row[0]),
                sku=_str(row[1]),
                product_cn_name=_str(row[2]),
                pr_date=_date(row[3]),
                pr_person=_str(row[4]),
                supplier=_str(row[5]),
                po_no=_str(row[6]),
                online_po_no=_str(row[7]),
                is_received=_str(row[8]),
                total_qty=_int(row[9]),
                total_boxes=_int(row[10]),
                product_label=_str(row[11]),
                carton_mark=_str(row[12]),
                warehouse_receipt=_str(row[14]),  # col 15 (0-indexed: 14)
                receiving_address=_str(row[16]),
                labeling_notes=_str(row[17]),
                logistics_provider=_str(row[18]),
                first_leg_tracking=_str(row[19]),
                total_boxes_2=_int(row[20]),
                length_cm=_dec(row[21]),
                width_cm=_dec(row[22]),
                height_cm=_dec(row[23]),
                gross_weight=_dec(row[24]),
                total_cbm=_dec(row[25]),
                density=_dec(row[26]),
                plan_no=_str(row[27]),
                ship_date=_date(row[28]),
                tracking_no=_str(row[29]),
                logistics_company=_str(row[30]),
                special_notes=_str(row[31]),
                previous_aftersales=_str(row[32]),
                qty_total_2=_int(row[33]),
                receiving_status=_str(row[34]),
                shipment_no=_str(row[35]),
            )
            db.add(item)
        result["shipment_count"] = ws.max_row - 1

    db.commit()
    return {"ok": True, **result}


# ============================================================
# Excel 导出
# ============================================================

@router.get("/export")
def export_all(db: Session = Depends(get_db)):
    """导出完整 ZIP — 结构对齐原始桌面 OZON直发信息 目录"""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="服务器未安装 openpyxl")

    import io, zipfile
    from openpyxl.styles import Font

    zip_buf = io.BytesIO()
    link_font = Font(color="0000FF", underline="single")

    LABEL_DIR = "OZON直发信息-FILE/SKU基础数据/标签文件"
    RECEIPT_DIR = "OZON直发信息-FILE/直发跟进表-N/入库清单"

    sku_count = 0
    ship_count = 0
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # ── Sheet 1: SKU基础数据 ──
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SKU基础数据"
        ws.append(["SKU", "产品名称", "供应商", "店铺", "标签文件", ""])
        items = db.query(OzonDirectSku).filter_by(is_deleted=False).order_by(OzonDirectSku.id).all()
        sku_count = len(items)
        for r, it in enumerate(items, 2):
            ws.cell(row=r, column=1, value=it.sku)
            ws.cell(row=r, column=2, value=it.product_name)
            ws.cell(row=r, column=3, value=it.supplier)
            ws.cell(row=r, column=4, value=it.store_name)
            cell = ws.cell(row=r, column=5, value=it.label_file)
            if it.label_file:
                cell.hyperlink = f"{LABEL_DIR}/{it.label_file}"
                cell.font = link_font

        # ── Sheet 2: 直发跟进表-N ──
        ws2 = wb.create_sheet("直发跟进表-N")
        headers = [
            "申购单号", "SKU", "产品中文名", "申购时间", "申购人员", "供应商",
            "采购单号", "网采单号", "是否收货上架", "总数", "总箱数",
            "产品标签", "外箱箱唛", "", "入库清单", "",
            "收货地址", "贴标发货说明", "物流商", "物流商头程单号", "总箱数.",
            "长", "宽", "高", "毛重", "总方数", "密度",
            "计划单号", "发货时间", "物流单号", "物流公司",
            "特殊情况备注", "上期售后情况", "总数.", "货物收货情况", "货件单号",
        ]
        for c, h in enumerate(headers, 1):
            ws2.cell(row=1, column=c, value=h)

        shipments = db.query(OzonDirectShipment).filter_by(is_deleted=False).order_by(OzonDirectShipment.id).all()
        ship_count = len(shipments)
        for r, it in enumerate(shipments, 2):
            ws2.cell(row=r, column=1, value=it.pr_no)
            ws2.cell(row=r, column=2, value=it.sku)
            ws2.cell(row=r, column=3, value=it.product_cn_name)
            ws2.cell(row=r, column=4, value=it.pr_date.strftime("%Y-%m-%d") if it.pr_date else None)
            ws2.cell(row=r, column=5, value=it.pr_person)
            ws2.cell(row=r, column=6, value=it.supplier)
            ws2.cell(row=r, column=7, value=it.po_no)
            ws2.cell(row=r, column=8, value=it.online_po_no)
            ws2.cell(row=r, column=9, value=it.is_received)
            ws2.cell(row=r, column=10, value=it.total_qty)
            ws2.cell(row=r, column=11, value=it.total_boxes)
            ws2.cell(row=r, column=12, value=it.product_label)
            ws2.cell(row=r, column=13, value=it.carton_mark)
            # col 14 — 空（原始 Excel 合并单元格）
            cell = ws2.cell(row=r, column=15, value=it.warehouse_receipt)
            if it.warehouse_receipt:
                cell.hyperlink = f"{RECEIPT_DIR}/{it.warehouse_receipt}"
                cell.font = link_font
            # col 16 — 空
            ws2.cell(row=r, column=17, value=it.receiving_address)
            ws2.cell(row=r, column=18, value=it.labeling_notes)
            ws2.cell(row=r, column=19, value=it.logistics_provider)
            ws2.cell(row=r, column=20, value=it.first_leg_tracking)
            ws2.cell(row=r, column=21, value=it.total_boxes_2)
            ws2.cell(row=r, column=22, value=float(it.length_cm) if it.length_cm else None)
            ws2.cell(row=r, column=23, value=float(it.width_cm) if it.width_cm else None)
            ws2.cell(row=r, column=24, value=float(it.height_cm) if it.height_cm else None)
            ws2.cell(row=r, column=25, value=float(it.gross_weight) if it.gross_weight else None)
            ws2.cell(row=r, column=26, value=float(it.total_cbm) if it.total_cbm else None)
            ws2.cell(row=r, column=27, value=float(it.density) if it.density else None)
            ws2.cell(row=r, column=28, value=it.plan_no)
            ws2.cell(row=r, column=29, value=it.ship_date.strftime("%Y-%m-%d") if it.ship_date else None)
            ws2.cell(row=r, column=30, value=it.tracking_no)
            ws2.cell(row=r, column=31, value=it.logistics_company)
            ws2.cell(row=r, column=32, value=it.special_notes)
            ws2.cell(row=r, column=33, value=it.previous_aftersales)
            ws2.cell(row=r, column=34, value=it.qty_total_2)
            ws2.cell(row=r, column=35, value=it.receiving_status)
            ws2.cell(row=r, column=36, value=it.shipment_no)

        xl_buf = io.BytesIO()
        wb.save(xl_buf)
        xl_buf.seek(0)
        zf.writestr("OZON直发信息.xlsx", xl_buf.read())

        # ── 标签文件 ──
        for f in db.query(OzonDirectFile).filter_by(source_table="sku").all():
            if f.file_data:
                zf.writestr(f"{LABEL_DIR}/{f.file_name}", f.file_data)

        # ── 入库清单 ──
        for f in db.query(OzonDirectFile).filter_by(source_table="shipment").all():
            if f.file_data:
                zf.writestr(f"{RECEIPT_DIR}/{f.file_name}", f.file_data)

    zip_buf.seek(0)
    log_operation("EXPORT", f"sku={sku_count} shipment={ship_count}")
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=OZON_export.zip"},
    )
